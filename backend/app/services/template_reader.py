from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

from app.config import settings
from app.schemas import SHEET_BY_METHODOLOGY, Methodology


@lru_cache(maxsize=8)
def _read_headers(template_path: str, sheet_name: str) -> tuple[str, ...]:
    workbook = load_workbook(filename=template_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Aba '{sheet_name}' não encontrada na planilha modelo.")

    sheet = workbook[sheet_name]
    headers: list[str] = []
    seen: set[str] = set()
    for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
        value = cell.value
        if value is None:
            continue
        text = str(value).strip()
        # Colunas de layout de entidades "x"/"y"/"n" se repetem no modelo; mantém únicas.
        if not text or text in seen:
            continue
        seen.add(text)
        headers.append(text)
    workbook.close()
    return tuple(headers)


def _template_for(methodology: Methodology) -> tuple[Path, str]:
    """SFP usa o modelo original; APF/NESMA usam a planilha 01_ENTRADA_PF_(APF)."""
    if methodology == Methodology.sfp:
        return settings.template_file, "TEMPLATE_PATH"
    return settings.template_apf_file, "TEMPLATE_APF_PATH"


def get_expected_fields(methodology: Methodology) -> list[str]:
    template, env_var = _template_for(methodology)
    if not Path(template).exists():
        raise FileNotFoundError(
            f"Planilha modelo não encontrada em '{template}'. "
            f"Ajuste {env_var} no arquivo .env."
        )
    sheet_name = SHEET_BY_METHODOLOGY[methodology]
    return list(_read_headers(str(template), sheet_name))
