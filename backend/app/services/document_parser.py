import io

from openpyxl import load_workbook

MAX_CHARS = 60_000


class UnsupportedDocumentError(Exception):
    pass


def _from_xlsx(data: bytes) -> str:
    try:
        workbook = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise UnsupportedDocumentError(
            "Não foi possível abrir o arquivo Excel. Envie o modelo "
            "'Padrão de Entrada PF.xlsx' preenchido."
        ) from exc

    parts: list[str] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [
            str(c).strip() if c is not None else "" for c in rows[0]
        ]
        data_rows = rows[1:]

        # Mantém apenas registros que tenham ao menos uma célula preenchida.
        records: list[list[tuple[str, str]]] = []
        for row in data_rows:
            pairs: list[tuple[str, str]] = []
            for idx, cell in enumerate(row):
                if cell is None:
                    continue
                value = str(cell).strip()
                if not value:
                    continue
                header = headers[idx] if idx < len(headers) and headers[idx] else f"Coluna {idx + 1}"
                pairs.append((header, value))
            if pairs:
                records.append(pairs)

        if not records:
            continue

        parts.append(f"## Aba: {sheet.title}")
        for i, pairs in enumerate(records, start=1):
            parts.append(f"Registro {i}:")
            for header, value in pairs:
                parts.append(f"  - {header}: {value}")

    workbook.close()
    return "\n".join(parts)


def extract_text(filename: str, data: bytes) -> str:
    name = (filename or "").lower()
    if not name.endswith(".xlsx"):
        raise UnsupportedDocumentError(
            "Formato não suportado. Envie apenas o modelo Excel (.xlsx) "
            "'Padrão de Entrada PF'."
        )

    text = _from_xlsx(data).strip()
    if not text:
        raise UnsupportedDocumentError(
            "O modelo Excel enviado está vazio. Preencha a aba da metodologia "
            "escolhida antes de anexar."
        )
    return text[:MAX_CHARS]
