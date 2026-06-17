"""
Gera o KIT DE TESTE do fluxo completo VMO.

Saída (pasta kit_teste/ na raiz do repo):
  01_ENTRADA_PF_(SFP).xlsx        -> solicitante anexa ao criar a solicitação
  02_PROPOSTA_OK.xlsx             -> fornecedor anexa: variação ~0%   (verde)
  03_PROPOSTA_ATENCAO.xlsx        -> fornecedor anexa: variação ~+24% (amarelo)
  04_PROPOSTA_DIVERGENTE.xlsx     -> fornecedor anexa: variação ~+90% (vermelho)

Uso: python -m scripts.gerar_kit_teste
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Estimativa-alvo que o motor de IA gera a partir da entrada abaixo.
# Medido em produção (temperature=0, resultado estável): a IA conta como AL
# TODAS as entidades internas acionadas (não só a "(nova)"), resultando em 77 PF.
PF_ALVO = 77

COR_AZUL     = "1E3A5F"
COR_AZUL_MED = "2D5F8A"
COR_AZUL_CLR = "D6E4F0"
COR_VERDE    = "1A7A4A"
COR_CINZA    = "F5F5F5"
COR_BRANCO   = "FFFFFF"
COR_AMARELO  = "FFF3CD"

OUT_DIR = Path(__file__).resolve().parents[2] / "kit_teste"
OUT_DIR.mkdir(exist_ok=True)


def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def hdr(ws, row, col, text, fg=COR_AZUL, bold=True, size=11, align="left", wrap=False, colspan=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold, color=COR_BRANCO if fg in (COR_AZUL, COR_AZUL_MED, COR_VERDE) else "000000",
                     size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin()
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + colspan - 1)
    return cell


def val(ws, row, col, value, bold=False, align="left", fmt=None, bg=COR_BRANCO, color="000000", colspan=1, wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, name="Calibri", size=10, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin()
    if fmt:
        cell.number_format = fmt
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + colspan - 1)
    return cell


def label(ws, row, col, text, colspan=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, name="Calibri", size=9, color="555555")
    cell.fill = PatternFill("solid", fgColor=COR_CINZA)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _thin()
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + colspan - 1)
    return cell


# ─────────────────────────────────────────────────────────────────────────────
# 1) ENTRADA PF (SFP) — o que o solicitante anexa
# ─────────────────────────────────────────────────────────────────────────────
def gerar_entrada():
    wb = Workbook()
    ws = wb.active
    ws.title = "SFP"

    colunas = [
        "Código Projeto/Iniciativa",
        "Nome do Projeto/Iniciativa",
        "Aplicação/Sistema",
        "Medição para contratação? (Sim/Não)",
        "Código da funcionalidade/serviço",
        "Nome da funcionalidade/serviço",
        "Comportamento atual",
        "Melhoria proposta",
        "Comportamento esperado após a melhoria",
        "Entidades de dados acionadas antes da melhoria",
        "Entidades de dados acionadas após a melhoria",
        "Possui entidade de dados nova ou alterada em razão da funcionalidade?",
        "Entidade de dados alterada é mantida nesta aplicação ou é externa?",
        "Nome da entidade de dados criada/alterada pela funcionalidade",
    ]
    for i, c in enumerate(colunas, 1):
        hdr(ws, 1, i, c, fg=COR_AZUL_MED, size=9, wrap=True)
        ws.column_dimensions[get_column_letter(i)].width = 26

    # 5 funcionalidades com mistura de entidades CRIADAS e ALTERADAS (mantidas nesta
    # aplicação). A coluna "Nome da entidade criada/alterada" é preenchida de forma
    # explícita em todos os registros => análise da IA fecha 14/14 campos.
    # Em projeto de DESENVOLVIMENTO, entidade alterada também conta como ADD,
    # então a estimativa permanece ~77 PF.
    funcs = [
        ("UC01", "Cadastrar pedido de compra",
         "O pedido é registrado manualmente em planilha pela equipe.",
         "Permitir o cadastro do pedido diretamente no portal.",
         "O usuário preenche o pedido e o sistema grava no banco.",
         "Pedido", "Pedido, ItemPedido", "Sim", "Mantida nesta aplicação",
         "ItemPedido (entidade nova, criada por esta funcionalidade)"),
        ("UC02", "Aprovar pedido de compra",
         "A aprovação é feita por e-mail, sem registro estruturado.",
         "Disponibilizar fluxo de aprovação com trilha de auditoria.",
         "O aprovador aceita ou recusa e o sistema registra a decisão.",
         "Pedido", "Pedido, Aprovacao", "Sim", "Mantida nesta aplicação",
         "Pedido (entidade existente, alterada por esta funcionalidade)"),
        ("UC03", "Consultar histórico de fornecedores",
         "A consulta é feita em sistema legado separado.",
         "Centralizar o histórico de fornecedores no portal.",
         "O usuário pesquisa e visualiza o histórico consolidado.",
         "Fornecedor", "Fornecedor, HistoricoFornecedor", "Sim", "Mantida nesta aplicação",
         "HistoricoFornecedor (entidade nova, criada por esta funcionalidade)"),
        ("UC04", "Emitir relatório de gastos por centro de custo",
         "Os relatórios são montados manualmente no fim do mês.",
         "Gerar relatório automático por centro de custo.",
         "O sistema consolida os gastos e gera o relatório em PDF.",
         "CentroCusto", "Pedido, CentroCusto", "Sim", "Mantida nesta aplicação",
         "CentroCusto (entidade existente, alterada por esta funcionalidade)"),
        ("UC05", "Notificar vencimento de contrato por e-mail",
         "Não há aviso automático de vencimento de contrato.",
         "Disparar e-mail automático antes do vencimento.",
         "O sistema consulta contratos a vencer e envia a notificação.",
         "Contrato", "Contrato, AgendaNotificacao", "Sim", "Mantida nesta aplicação",
         "AgendaNotificacao (entidade nova, criada por esta funcionalidade)"),
    ]

    for r, f in enumerate(funcs, start=2):
        cod, nome, atual, melhoria, esperado, ent_antes, ent_depois, possui, mantida, entidade = f
        linha = [
            "PRJ-2026-VMO", "Modernização de Compras", "Portal de Compras", "Sim",
            cod, nome, atual, melhoria, esperado, ent_antes, ent_depois, possui, mantida, entidade,
        ]
        for i, v in enumerate(linha, 1):
            val(ws, r, i, v, bg=COR_AZUL_CLR if r % 2 == 0 else COR_BRANCO, align="left")

    out = OUT_DIR / "01_ENTRADA_PF_(SFP).xlsx"
    wb.save(out)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# 2) PROPOSTAS — o que o fornecedor anexa
# ─────────────────────────────────────────────────────────────────────────────
def gerar_proposta(total_pf: int, cenario: str, nome_arquivo: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Proposta Comercial"
    ws.sheet_view.showGridLines = False

    for i, w in enumerate([6, 34, 12, 12, 12, 12, 16, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[2].height = 34
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = "PROPOSTA COMERCIAL"
    c.font = Font(bold=True, color=COR_BRANCO, size=18, name="Calibri")
    c.fill = PatternFill("solid", fgColor=COR_AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:H3")
    c = ws["A3"]
    c.value = f"Modernização de Compras — Cenário de teste: {cenario}"
    c.font = Font(color=COR_BRANCO, size=11, name="Calibri")
    c.fill = PatternFill("solid", fgColor=COR_AZUL_MED)
    c.alignment = Alignment(horizontal="center", vertical="center")

    r = 5
    ws.merge_cells(f"A{r}:H{r}")
    hdr(ws, r, 1, "  1. IDENTIFICAÇÃO", colspan=8)
    r = 6
    label(ws, r, 1, "Fornecedor"); val(ws, r, 2, "TechSoft Soluções Ltda.", bold=True, colspan=3)
    label(ws, r, 5, "CNPJ"); val(ws, r, 6, "12.345.678/0001-90", colspan=3)
    r = 7
    label(ws, r, 1, "Contato"); val(ws, r, 2, "João Silva", colspan=3)
    label(ws, r, 5, "Validade"); val(ws, r, 6, "30 dias", colspan=3)

    r = 9
    ws.merge_cells(f"A{r}:H{r}")
    hdr(ws, r, 1, "  2. ESCOPO TÉCNICO — CONTAGEM SFP 2.2", colspan=8)
    r = 10
    for col, txt in enumerate(["#", "Funcionalidade / Componente", "Tipo", "Operação",
                               "PF Unit.", "Qtd", "PF Total", "Observação"], 1):
        hdr(ws, r, col, txt, fg=COR_AZUL_MED, size=9)

    # Distribui o total alvo em funcionalidades plausíveis
    itens = [
        ("Cadastro de pedido de compra", "EQ", "ADD"),
        ("Fluxo de aprovação de pedido", "EQ", "ADD"),
        ("Consulta de histórico de fornecedores", "SE", "INF"),
        ("Relatório de gastos por centro de custo", "SE", "AGL"),
        ("Notificação de vencimento de contrato", "EQ", "ADD"),
        ("Autenticação e controle de acesso", "EQ", "ADD"),
        ("Log de auditoria (LGPD)", "EE", "ADD"),
        ("Testes automatizados e deploy CI/CD", "EQ", "ALT"),
    ]
    n = len(itens)
    base = total_pf // n
    resto = total_pf - base * n
    pfs = [base + (1 if i < resto else 0) for i in range(n)]

    for i, ((desc, tipo, op), pf_t) in enumerate(zip(itens, pfs)):
        r += 1
        bg = COR_AZUL_CLR if i % 2 == 0 else COR_BRANCO
        val(ws, r, 1, i + 1, align="center", bg=bg)
        val(ws, r, 2, desc, bg=bg)
        val(ws, r, 3, tipo, align="center", bg=bg)
        val(ws, r, 4, op, align="center", bg=bg)
        val(ws, r, 5, pf_t, align="center", bg=bg, fmt="0")
        val(ws, r, 6, 1, align="center", bg=bg, fmt="0")
        val(ws, r, 7, pf_t, align="center", bg=bg, bold=True, fmt="0")
        val(ws, r, 8, "—", bg=bg)

    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    hdr(ws, r, 1, "TOTAL DE PONTOS DE FUNÇÃO (PF BRUTOS)", fg=COR_VERDE, colspan=6)
    val(ws, r, 7, total_pf, bold=True, align="center", bg=COR_VERDE, color=COR_BRANCO, fmt="0")
    ws.cell(r, 8).fill = PatternFill("solid", fgColor=COR_VERDE)

    # Precificação
    valor_pf = 820.00
    valor_total = total_pf * valor_pf
    r += 2
    ws.merge_cells(f"A{r}:H{r}")
    hdr(ws, r, 1, "  3. PRECIFICAÇÃO", colspan=8)
    r += 1
    label(ws, r, 1, "Valor R$/PF", colspan=2)
    val(ws, r, 3, valor_pf, bold=True, fmt='R$ #,##0.00', align="right", bg=COR_AMARELO)
    label(ws, r, 4, "Total PF"); val(ws, r, 5, total_pf, bold=True, fmt="0", align="center")
    label(ws, r, 6, "Valor Total"); val(ws, r, 7, valor_total, bold=True, fmt='R$ #,##0.00', align="right", bg=COR_AMARELO, colspan=2)
    r += 1
    label(ws, r, 1, "Prazo de Entrega", colspan=2); val(ws, r, 3, "2026-09-30")
    label(ws, r, 4, "Modalidade"); val(ws, r, 5, "Ponto de Função (SFP 2.2)", colspan=4)

    # Rodapé lido pelo Motor De-Para (regex "Total PF: N")
    r += 2
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(r, 1)
    c.value = f"Resumo para Motor APF  |  Total PF: {total_pf}  |  R$/PF: R$ {valor_pf:,.2f}  |  Valor Total: R$ {valor_total:,.2f}"
    c.font = Font(bold=True, color=COR_BRANCO, size=9, name="Calibri")
    c.fill = PatternFill("solid", fgColor=COR_VERDE)
    c.alignment = Alignment(horizontal="center", vertical="center")

    out = OUT_DIR / nome_arquivo
    wb.save(out)
    return out


if __name__ == "__main__":
    e = gerar_entrada()
    print(f"Entrada gerada     : {e.name}  (estimativa-alvo ~ {PF_ALVO} PF)")

    cenarios = [
        (PF_ALVO,                 "OK (dentro do esperado)",   "02_PROPOSTA_OK.xlsx"),
        (round(PF_ALVO * 1.24),   "ATENÇÃO (variação média)",  "03_PROPOSTA_ATENCAO.xlsx"),
        (round(PF_ALVO * 1.90),   "DIVERGENTE (muito acima)",  "04_PROPOSTA_DIVERGENTE.xlsx"),
    ]
    for total, cen, fname in cenarios:
        p = gerar_proposta(total, cen, fname)
        var = (total - PF_ALVO) / PF_ALVO * 100
        print(f"Proposta gerada    : {p.name}  (Total PF: {total} · variação {var:+.0f}%)")

    print(f"\nArquivos em: {OUT_DIR}")
