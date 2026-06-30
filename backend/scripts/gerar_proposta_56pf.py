"""
Gera proposta com Total PF = 56 (igual à estimativa inicial) para teste de De-Para.
Uso: python -m scripts.gerar_proposta_56pf
Saída: proposta_techsoft_56pf.xlsx (na raiz do backend)
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

COR_AZUL     = "1E3A5F"
COR_AZUL_MED = "2D5F8A"
COR_AZUL_CLR = "D6E4F0"
COR_VERDE    = "1A7A4A"
COR_CINZA    = "F5F5F5"
COR_BRANCO   = "FFFFFF"
COR_AMARELO  = "FFF3CD"


def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def hdr(ws, row, col, text, fg=COR_AZUL, bold=True, size=11, align="left", colspan=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold, color=COR_BRANCO if fg in (COR_AZUL, COR_AZUL_MED, COR_VERDE) else "000000",
                     size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _thin()
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + colspan - 1)
    return cell


def val(ws, row, col, value, bold=False, align="left", fmt=None, bg=COR_BRANCO, color="000000", colspan=1):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, name="Calibri", size=10, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _thin()
    if fmt:
        cell.number_format = fmt
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + colspan - 1)
    return cell


def label(ws, row, col, text, colspan=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, name="Calibri", size=9, color="555555")
    cell.fill = PatternFill("solid", fgColor=COR_CINZA)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _thin()
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + colspan - 1)
    return cell


wb = Workbook()
ws = wb.active
ws.title = "Proposta Comercial"
ws.sheet_view.showGridLines = False

for i, w in enumerate([6, 30, 18, 14, 14, 14, 16, 20], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Cabeçalho ────────────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 36
ws.row_dimensions[3].height = 22

ws.merge_cells("A2:H2")
c = ws["A2"]
c.value = "PROPOSTA COMERCIAL"
c.font = Font(bold=True, color=COR_BRANCO, size=18, name="Calibri")
c.fill = PatternFill("solid", fgColor=COR_AZUL)
c.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A3:H3")
c = ws["A3"]
c.value = "Contrato de Serviços de TI – Modalidade Ponto de Função (SFP 2.2)"
c.font = Font(bold=False, color=COR_BRANCO, size=11, name="Calibri")
c.fill = PatternFill("solid", fgColor=COR_AZUL_MED)
c.alignment = Alignment(horizontal="center", vertical="center")

# ── 1. Identificação ─────────────────────────────────────────────────────────
r = 5
ws.merge_cells(f"A{r}:H{r}")
hdr(ws, r, 1, "  1. IDENTIFICAÇÃO DA SOLICITAÇÃO", colspan=8)

r = 6
label(ws, r, 1, "Nº Solicitação"); val(ws, r, 2, "Proposta TechSoft", bold=True)
label(ws, r, 3, "Data Proposta");  val(ws, r, 4, "17/06/2026")
label(ws, r, 5, "Validade");       val(ws, r, 6, "30 dias")
label(ws, r, 7, "Revisão");        val(ws, r, 8, "1.0")

r = 7
label(ws, r, 1, "Fornecedor"); val(ws, r, 2, "TechSoft Soluções Ltda.", bold=True, colspan=3)
label(ws, r, 5, "CNPJ");       val(ws, r, 6, "12.345.678/0001-90", colspan=3)

# ── 2. Escopo Técnico ─────────────────────────────────────────────────────────
r = 9
ws.merge_cells(f"A{r}:H{r}")
hdr(ws, r, 1, "  2. ESCOPO TÉCNICO – CONTAGEM SFP 2.2", colspan=8)

r = 10
for col, txt in enumerate(["#", "Funcionalidade / Componente", "Tipo",
                            "Operação", "PF Unit.", "Qtd", "PF Total", "Observação"], 1):
    hdr(ws, r, col, txt, fg=COR_AZUL_MED)

# Funções projetadas para somar exatamente 56 PF
funcoes = [
    (1,  "Cadastro e gestão de contratos",            "EQ", "ADD", 5, 2, 10, "CRUD completo"),
    (2,  "Consulta e filtros avançados",              "SE", "INF", 3, 2,  6, "Listagem + exportação"),
    (3,  "Upload e gestão de documentos",             "EQ", "ADD", 4, 2,  8, "Supabase Storage"),
    (4,  "Dashboard e indicadores gerenciais",        "SE", "AGL", 4, 2,  8, "Gráficos + KPIs"),
    (5,  "Autenticação e controle de acesso (RBAC)",  "EQ", "ADD", 5, 1,  5, "JWT + refresh token"),
    (6,  "Integração com sistema legado via API",     "EQ", "ADD", 7, 1,  7, "REST bidirecional"),
    (7,  "Notificações automáticas por e-mail",       "EQ", "ADD", 5, 1,  5, "SendGrid"),
    (8,  "Log de auditoria e rastreabilidade",        "EE", "ADD", 3, 1,  3, "LGPD compliance"),
    (9,  "Testes automatizados (unitário + e2e)",     "EQ", "ADD", 4, 1,  4, "Cobertura ≥80%"),
]
# Total: 10+6+8+8+5+7+5+3+4 = 56

total_pf = 0
for i, (num, desc, tipo, op, pf_u, qtd, pf_t, obs) in enumerate(funcoes):
    r += 1
    bg = COR_AZUL_CLR if i % 2 == 0 else COR_BRANCO
    val(ws, r, 1, num,  align="center", bg=bg)
    val(ws, r, 2, desc, bg=bg)
    val(ws, r, 3, tipo, align="center", bg=bg)
    val(ws, r, 4, op,   align="center", bg=bg)
    val(ws, r, 5, pf_u, align="center", bg=bg, fmt="0")
    val(ws, r, 6, qtd,  align="center", bg=bg, fmt="0")
    val(ws, r, 7, pf_t, align="center", bg=bg, bold=True, fmt="0")
    val(ws, r, 8, obs,  bg=bg)
    total_pf += pf_t

r += 1
ws.merge_cells(f"A{r}:F{r}")
hdr(ws, r, 1, "TOTAL DE PONTOS DE FUNÇÃO (PF BRUTOS)", fg=COR_VERDE, colspan=6)
val(ws, r, 7, total_pf, bold=True, align="center", bg=COR_VERDE, color=COR_BRANCO, fmt="0")
ws.cell(r, 8).fill = PatternFill("solid", fgColor=COR_VERDE)

# ── 3. Precificação ───────────────────────────────────────────────────────────
r += 2
ws.merge_cells(f"A{r}:H{r}")
hdr(ws, r, 1, "  3. PRECIFICAÇÃO", colspan=8)

valor_pf    = 1_380.00
valor_total = total_pf * valor_pf

r += 1
label(ws, r, 1, "Valor R$/PF proposto", colspan=2)
val(ws,   r, 3, valor_pf,    bold=True, fmt='R$ #,##0.00', align="right", bg=COR_AMARELO)
label(ws, r, 4, "Total PF")
val(ws,   r, 5, total_pf,    bold=True, fmt="0", align="center")
label(ws, r, 6, "Valor Total Estimado")
val(ws,   r, 7, valor_total, bold=True, fmt='R$ #,##0.00', align="right", bg=COR_AMARELO, colspan=2)

r += 1
label(ws, r, 1, "Prazo de Entrega", colspan=2); val(ws, r, 3, "2026-08-15")
label(ws, r, 4, "Modalidade"); val(ws, r, 5, "Ponto de Função (SFP 2.2)", colspan=4)

r += 1
label(ws, r, 1, "Forma de Pagamento", colspan=2)
val(ws, r, 3, "30/60/90 dias após aceite formal", colspan=6)

# ── Rodapé resumo motor APF ───────────────────────────────────────────────────
r += 2
ws.merge_cells(f"A{r}:H{r}")
c = ws.cell(r, 1)
c.value = (f"Resumo para Motor APF  |  Total PF: {total_pf}"
           f"  |  R$/PF: R$ {valor_pf:,.2f}  |  Valor Total: R$ {valor_total:,.2f}")
c.font = Font(bold=True, color=COR_BRANCO, size=9, name="Calibri")
c.fill = PatternFill("solid", fgColor=COR_VERDE)
c.alignment = Alignment(horizontal="center", vertical="center")

out = Path(__file__).parent.parent / "proposta_techsoft_56pf.xlsx"
wb.save(out)
print(f"Planilha gerada: {out}")
print(f"  Total PF   : {total_pf}")
print(f"  R$/PF      : R$ {valor_pf:,.2f}")
print(f"  Valor Total: R$ {valor_total:,.2f}")
