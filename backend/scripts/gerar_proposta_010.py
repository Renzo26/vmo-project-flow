"""
Gera a planilha de Proposta Comercial para VMO-2026-010.
Uso: python -m scripts.gerar_proposta_010
Saída: proposta_techsoft_VMO-2026-010.xlsx (na raiz do backend)
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

COR_AZUL     = "1E3A5F"
COR_AZUL_MED = "2D5F8A"
COR_AZUL_CLR = "D6E4F0"
COR_VERDE    = "1A7A4A"
COR_CINZA    = "F5F5F5"
COR_BRANCO   = "FFFFFF"
COR_AMARELO  = "FFF3CD"


def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def hdr(ws, row, col, text, fg=COR_AZUL, bg=COR_BRANCO, bold=True, size=11,
        align="left", wrap=False, colspan=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold, color=COR_BRANCO if fg in (COR_AZUL, COR_AZUL_MED, COR_VERDE)
                     else "000000", size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin()
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + colspan - 1)
    return cell


def val(ws, row, col, value, bold=False, align="left", fmt=None,
        bg=COR_BRANCO, color="000000", colspan=1):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, name="Calibri", size=10, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _thin()
    if fmt:
        cell.number_format = fmt
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + colspan - 1)
    return cell


def label(ws, row, col, text, colspan=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, name="Calibri", size=9, color="555555")
    cell.fill = PatternFill("solid", fgColor=COR_CINZA)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _thin()
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + colspan - 1)
    return cell


wb = Workbook()
ws = wb.active
ws.title = "Proposta Comercial"
ws.sheet_view.showGridLines = False
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1

for i, w in enumerate([6, 30, 18, 14, 14, 14, 16, 20], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Cabeçalho ────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 14
ws.row_dimensions[2].height = 36
ws.row_dimensions[3].height = 22

ws.merge_cells("A2:H2")
c = ws["A2"]
c.value = "PROPOSTA COMERCIAL"
c.font = Font(bold=True, color=COR_BRANCO, size=18, name="Calibri")
c.fill = PatternFill("solid", fgColor=COR_AZUL)
c.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A3:H3")
c = ws["A3"]
c.value = "Contrato de Serviços de TI – Modalidade Ponto de Função (SFP 2.2)"
c.font = Font(bold=False, color=COR_BRANCO, size=11, name="Calibri")
c.fill = PatternFill("solid", fgColor=COR_AZUL_MED)
c.alignment = Alignment(horizontal="center", vertical="center")

# ── 1. Identificação ─────────────────────────────────────────────────────────
r = 5
ws.row_dimensions[r].height = 16
ws.merge_cells(f"A{r}:H{r}")
hdr(ws, r, 1, "  1. IDENTIFICAÇÃO DA SOLICITAÇÃO", fg=COR_AZUL, colspan=8)

r = 6; ws.row_dimensions[r].height = 20
label(ws, r, 1, "Nº Solicitação")
val(ws,   r, 2, "VMO-2026-010", bold=True)
label(ws, r, 3, "Data Proposta")
val(ws,   r, 4, "16/06/2026")
label(ws, r, 5, "Validade")
val(ws,   r, 6, "30 dias")
label(ws, r, 7, "Revisão")
val(ws,   r, 8, "1.0")

r = 7; ws.row_dimensions[r].height = 20
label(ws, r, 1, "Título")
val(ws,   r, 2, "Portal de Autoatendimento – Gov.Br / CRM", bold=True, colspan=3)
label(ws, r, 5, "Tipo Projeto")
val(ws,   r, 6, "Desenvolvimento de Software", colspan=3)

r = 8; ws.row_dimensions[r].height = 20
label(ws, r, 1, "Fornecedor")
val(ws,   r, 2, "TechSoft Soluções Ltda.", bold=True, colspan=3)
label(ws, r, 5, "CNPJ")
val(ws,   r, 6, "12.345.678/0001-90", colspan=3)

r = 9; ws.row_dimensions[r].height = 20
label(ws, r, 1, "Contato")
val(ws,   r, 2, "João Silva – joao.silva@techsoft.com.br", colspan=3)
label(ws, r, 5, "Telefone")
val(ws,   r, 6, "(11) 98765-4321", colspan=3)

# ── 2. Escopo Técnico ─────────────────────────────────────────────────────────
r = 11; ws.row_dimensions[r].height = 16
ws.merge_cells(f"A{r}:H{r}")
hdr(ws, r, 1, "  2. ESCOPO TÉCNICO – CONTAGEM SFP 2.2", fg=COR_AZUL, colspan=8)

r = 12; ws.row_dimensions[r].height = 20
for col, txt in enumerate(["#", "Funcionalidade / Componente", "Tipo",
                            "Operação", "PF Unit.", "Qtd", "PF Total", "Observação"], 1):
    hdr(ws, r, col, txt, fg=COR_AZUL_MED)

funcoes = [
    (1,  "Portal de login unificado (Gov.Br SSO)",         "EQ", "INF", 5, 1,  5, "Integração Gov.Br"),
    (2,  "Cadastro e atualização de perfil de usuário",    "EQ", "AGL", 4, 2,  8, "LGPD compliant"),
    (3,  "Painel de solicitações do cidadão",              "SE", "INF", 3, 2,  6, "Listagem + filtros"),
    (4,  "Abertura de chamado via portal",                 "EQ", "INF", 5, 3, 15, "Com upload de anexos"),
    (5,  "Consulta de status de chamado em tempo real",    "SE", "INF", 3, 2,  6, "Integração CRM"),
    (6,  "Notificações por e-mail e push",                 "EQ", "AGL", 4, 1,  4, "SendGrid / FCM"),
    (7,  "Integração bidirecional com CRM legado",         "EQ", "INF", 7, 1,  7, "REST + webhook"),
    (8,  "Relatório gerencial de atendimentos",            "SE", "AGL", 4, 1,  4, "Exportação PDF/XLS"),
    (9,  "Autenticação e controle de acesso (RBAC)",       "EQ", "INF", 5, 1,  5, "JWT + refresh token"),
    (10, "Log de auditoria e rastreabilidade",             "EE", "INF", 3, 1,  3, "LGPD / SOX"),
    (11, "Testes automatizados (unitário + integração)",   "EQ", "ALT", 7, 1,  7, "Cobertura ≥80%"),
    (12, "Deploy em nuvem com CI/CD (GitHub Actions)",     "EQ", "INF", 5, 1,  5, "Docker + EasyPanel"),
]

total_pf = 0
for i, (num, desc, tipo, op, pf_u, qtd, pf_t, obs) in enumerate(funcoes):
    r += 1
    ws.row_dimensions[r].height = 20
    bg = COR_AZUL_CLR if i % 2 == 0 else COR_BRANCO
    val(ws, r, 1, num,   align="center", bg=bg)
    val(ws, r, 2, desc,  bg=bg)
    val(ws, r, 3, tipo,  align="center", bg=bg)
    val(ws, r, 4, op,    align="center", bg=bg)
    val(ws, r, 5, pf_u,  align="center", bg=bg, fmt="0")
    val(ws, r, 6, qtd,   align="center", bg=bg, fmt="0")
    val(ws, r, 7, pf_t,  align="center", bg=bg, bold=True, fmt="0")
    val(ws, r, 8, obs,   bg=bg)
    total_pf += pf_t

r += 1; ws.row_dimensions[r].height = 22
ws.merge_cells(f"A{r}:F{r}")
hdr(ws, r, 1, "TOTAL DE PONTOS DE FUNÇÃO (PF BRUTOS)", fg=COR_VERDE, colspan=6)
val(ws, r, 7, total_pf, bold=True, align="center", bg=COR_VERDE, color=COR_BRANCO, fmt="0")
ws.cell(r, 8).fill = PatternFill("solid", fgColor=COR_VERDE)

# ── 3. Precificação ───────────────────────────────────────────────────────────
r += 2; ws.row_dimensions[r].height = 16
ws.merge_cells(f"A{r}:H{r}")
hdr(ws, r, 1, "  3. PRECIFICAÇÃO", fg=COR_AZUL, colspan=8)

valor_pf    = 1_380.00
valor_total = total_pf * valor_pf

r += 1; ws.row_dimensions[r].height = 20
label(ws, r, 1, "Valor R$/PF proposto", colspan=2)
val(ws,   r, 3, valor_pf,    bold=True, fmt='R$ #,##0.00', align="right", bg=COR_AMARELO)
label(ws, r, 4, "Total PF")
val(ws,   r, 5, total_pf,    bold=True, fmt="0", align="center")
label(ws, r, 6, "Valor Total Estimado")
val(ws,   r, 7, valor_total, bold=True, fmt='R$ #,##0.00', align="right",
    bg=COR_AMARELO, colspan=2)

r += 1; ws.row_dimensions[r].height = 20
label(ws, r, 1, "Prazo de Entrega", colspan=2)
val(ws,   r, 3, "2026-08-15")
label(ws, r, 4, "Modalidade")
val(ws,   r, 5, "Ponto de Função (SFP 2.2)", colspan=4)

r += 1; ws.row_dimensions[r].height = 20
label(ws, r, 1, "Forma de Pagamento", colspan=2)
val(ws,   r, 3, "30/60/90 dias após aceite formal", colspan=6)

r += 1; ws.row_dimensions[r].height = 20
label(ws, r, 1, "Reajuste", colspan=2)
val(ws,   r, 3, "INPC/IBGE, após 12 meses de contrato", colspan=6)

# ── 4. Premissas ──────────────────────────────────────────────────────────────
r += 2; ws.row_dimensions[r].height = 16
ws.merge_cells(f"A{r}:H{r}")
hdr(ws, r, 1, "  4. PREMISSAS E CONDIÇÕES", fg=COR_AZUL, colspan=8)

premissas = [
    "O escopo contempla todas as funcionalidades listadas na seção 2.",
    "Não estão inclusos custos de infraestrutura (servidores, licenças de terceiros).",
    "Mudanças de escopo serão tratadas via Termo Aditivo e nova contagem PF.",
    "O contratante disponibilizará ambiente de homologação com dados representativos.",
    "SLA de suporte pós-entrega: 90 dias (correção de bugs sem custo adicional).",
    "Credenciais de integração Gov.Br e CRM serão fornecidas pela contratante.",
]

for i, texto in enumerate(premissas, 1):
    r += 1; ws.row_dimensions[r].height = 18
    bg = COR_AZUL_CLR if i % 2 == 0 else COR_BRANCO
    val(ws, r, 1, i,     align="center", bg=bg)
    val(ws, r, 2, texto, bg=bg, colspan=7)

# ── 5. Assinatura ─────────────────────────────────────────────────────────────
r += 2; ws.row_dimensions[r].height = 16
ws.merge_cells(f"A{r}:H{r}")
hdr(ws, r, 1, "  5. ASSINATURA", fg=COR_AZUL, colspan=8)

r += 2; ws.row_dimensions[r].height = 18
val(ws, r, 2, "TechSoft Soluções Ltda.",        bold=True, colspan=3)
val(ws, r, 5, "",                               colspan=4)

r += 1; ws.row_dimensions[r].height = 18
val(ws, r, 2, "João Silva – Diretor Comercial", colspan=3)
val(ws, r, 5, "Braesp – Área Contratante",      colspan=4)

r += 1; ws.row_dimensions[r].height = 18
label(ws, r, 2, "____________________________", colspan=3)
label(ws, r, 5, "____________________________", colspan=4)

r += 1; ws.row_dimensions[r].height = 18
val(ws, r, 2, "Local/Data: ________________  2026", colspan=3)
val(ws, r, 5, "Local/Data: ________________  2026", colspan=4)

# ── Rodapé resumo motor APF ───────────────────────────────────────────────────
r += 2; ws.row_dimensions[r].height = 16
ws.merge_cells(f"A{r}:H{r}")
c = ws.cell(r, 1)
c.value = (f"Resumo para Motor APF  |  Total PF: {total_pf}"
           f"  |  R$/PF: R$ {valor_pf:,.2f}  |  Valor Total: R$ {valor_total:,.2f}")
c.font = Font(bold=True, color=COR_BRANCO, size=9, name="Calibri")
c.fill = PatternFill("solid", fgColor=COR_VERDE)
c.alignment = Alignment(horizontal="center", vertical="center")

out = Path(__file__).parent.parent / "proposta_techsoft_VMO-2026-010.xlsx"
wb.save(out)
print(f"Planilha gerada: {out}")
print(f"  Total PF   : {total_pf}")
print(f"  R$/PF      : R$ {valor_pf:,.2f}")
print(f"  Valor Total: R$ {valor_total:,.2f}")
